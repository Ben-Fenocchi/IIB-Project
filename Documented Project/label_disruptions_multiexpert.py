import os
import json
import time
import random
from typing import Dict, Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from openai import OpenAI
from google import genai


# ============================
# Paths
# ============================
INPUT_XLSX  = os.path.join("data", "interim", "disruption_master_10k_multiexpert.xlsx")
OUTPUT_XLSX = os.path.join("data", "interim", "disruption_master_10k_multiexpert_labelled.xlsx")

# ============================
# Models
# ============================
OPENAI_MODEL = "gpt-4.1-mini"
GEMINI_MODEL = "gemini-3-flash-preview"

# ============================
# Runtime controls
# ============================
SAVE_EVERY = 25
SLEEP_SECONDS = 0.15
MAX_RETRIES = 4

# ============================
# Disruption types (must match columns)
# ============================
TYPES = [
    "flood",
    "drought",
    "cyclone_huricane",
    "extreme_heat",
    "landslide",
    "earthquake",
    "mine_accident",
    "labour_strike",
    "protests",
    "trade_embargo",
    "country_relations",
    "tariffs",
]

# ============================
# Prompt (NO rationale)
# ============================
SYSTEM_RUBRIC = """
You are labeling news articles for supply-chain disruption signals.

Return a JSON object with ONE 0/1 label for EACH disruption type.

Rules:
- Use 1 ONLY if the article describes that disruption type as an actual event
  affecting operations, production, logistics, trade, or access.
- Use 0 for commentary, background, forecasts, financial reporting, policy discussion
  with no operational impact, historical reference, or unrelated content.
- If unclear or ambiguous, default to 0.

Be conservative: false positives are costly downstream.

Output MUST be valid JSON and match the schema exactly.
""".strip()


def make_user_text(url: str, title: str, meta: str) -> str:
    return (
        f"URL: {url}\n"
        f"TITLE: {title}\n"
        f"META: {meta}\n\n"
        f"Disruption types: {TYPES}\n"
        f"Return JSON with keys: {TYPES}\n"
    )


# ============================
# OpenAI Structured Output schema (no rationale)
# ============================
OPENAI_JSON_SCHEMA = {
    "name": "disruption_multilabel",
    "schema": {
        "type": "object",
        "properties": {
            **{t: {"type": "integer", "enum": [0, 1]} for t in TYPES}
        },
        "required": TYPES,
        "additionalProperties": False,
    },
    "strict": True,
}


# ============================
# Helpers
# ============================
def _coerce_label(v: Any) -> int:
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)) and v in (0, 1):
        return int(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("1", "true", "yes"):
            return 1
        if s in ("0", "false", "no"):
            return 0
    raise ValueError(f"Invalid label value: {v!r}")


def validate_payload(payload: Dict[str, Any]) -> Dict[str, int]:
    return {t: _coerce_label(payload[t]) for t in TYPES}


def backoff_sleep(attempt: int) -> None:
    time.sleep((2 ** attempt) * 0.6 + random.random() * 0.3)


def strip_fences(text: str) -> str:
    t = (text or "").strip()
    if t.startswith("```"):
        parts = t.split("```")
        if len(parts) >= 3:
            return parts[1].strip()
    return t


# ============================
# API calls
# ============================
def call_openai_labels(client: OpenAI, user_text: str) -> Dict[str, int]:
    resp = client.responses.create(
        model=OPENAI_MODEL,
        input=[
            {"role": "system", "content": SYSTEM_RUBRIC},
            {"role": "user", "content": user_text},
        ],
        text={"format": {"type": "json_schema", "json_schema": OPENAI_JSON_SCHEMA}},
        temperature=0,
    )
    return validate_payload(json.loads(resp.output_text))


def call_gemini_labels(client: genai.Client, user_text: str) -> Dict[str, int]:
    prompt = SYSTEM_RUBRIC + "\n\n" + user_text + "\nReturn ONLY JSON."
    resp = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=prompt,
    )
    return validate_payload(json.loads(strip_fences(resp.text)))


# ============================
# Excel helpers
# ============================
def header_map(ws: Worksheet) -> Dict[str, int]:
    return {
        ws.cell(row=1, column=c).value.strip(): c
        for c in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row=1, column=c).value, str)
    }


def get_cell(ws: Worksheet, row: int, col: str, hm: Dict[str, int]):
    return ws.cell(row=row, column=hm[col]).value


def set_cell(ws: Worksheet, row: int, col: str, hm: Dict[str, int], val):
    ws.cell(row=row, column=hm[col], value=val)


def any_missing(ws: Worksheet, row: int, hm: Dict[str, int], prefix: str) -> bool:
    return any(
        get_cell(ws, row, f"{prefix}{t}", hm) in (None, "")
        for t in TYPES
    )


# ============================
# Main
# ============================
def main():
    load_dotenv()

    openai_key = os.getenv("OPENAI_PROJECT_KEY") or os.getenv("OPENAI_ADMIN_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY")

    if not openai_key:
        raise RuntimeError("Missing OPENAI_PROJECT_KEY or OPENAI_ADMIN_KEY")
    if not gemini_key:
        raise RuntimeError("Missing GEMINI_API_KEY")

    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)

    openai_client = OpenAI(api_key=openai_key)
    gemini_client = genai.Client(api_key=gemini_key)

    wb = load_workbook(INPUT_XLSX)
    ws = wb["data"]
    hm = header_map(ws)

    for t in TYPES:
        for p in ("chatgpt_", "gemini_"):
            if f"{p}{t}" not in hm:
                raise RuntimeError(f"Missing column: {p}{t}")

    processed = skipped_gold = skipped_done = 0

    for r in range(2, ws.max_row + 1):
        if (get_cell(ws, r, "row_origin", hm) or "") == "gold_manual":
            skipped_gold += 1
            continue

        need_cg = any_missing(ws, r, hm, "chatgpt_")
        need_gm = any_missing(ws, r, hm, "gemini_")
        if not need_cg and not need_gm:
            skipped_done += 1
            continue

        url   = str(get_cell(ws, r, "url_normalized", hm) or "")
        title = str(get_cell(ws, r, "title", hm) or "")
        meta  = str(get_cell(ws, r, "meta_description", hm) or "")
        if not (url or title or meta):
            continue

        user_text = make_user_text(url, title, meta)

        if need_cg:
            for a in range(MAX_RETRIES):
                try:
                    out = call_openai_labels(openai_client, user_text)
                    for t in TYPES:
                        set_cell(ws, r, f"chatgpt_{t}", hm, out[t])
                    break
                except Exception:
                    if a == MAX_RETRIES - 1:
                        pass
                    else:
                        backoff_sleep(a)

        if need_gm:
            for a in range(MAX_RETRIES):
                try:
                    out = call_gemini_labels(gemini_client, user_text)
                    for t in TYPES:
                        set_cell(ws, r, f"gemini_{t}", hm, out[t])
                    break
                except Exception:
                    if a == MAX_RETRIES - 1:
                        pass
                    else:
                        backoff_sleep(a)

        processed += 1
        time.sleep(SLEEP_SECONDS)

        if processed % SAVE_EVERY == 0:
            wb.save(OUTPUT_XLSX)
            print(f"Checkpoint: {processed} labelled")

    wb.save(OUTPUT_XLSX)
    print("Done.")
    print(f"Labelled: {processed}")
    print(f"Skipped gold: {skipped_gold}")
    print(f"Skipped done: {skipped_done}")
    print(f"Saved to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
